#!/usr/bin/env python3
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill


def estilizar(sheet, lista, estilo, tipo):
    """
    Aplica um estilo a uma lista de células.
    sheet : worksheet do openpyxl
    lista : ['A1', 'B1', ...]
    estilo: objeto de estilo (Font, PatternFill, Border, etc.)
    tipo  : 'font', 'fill', 'alignment', 'border'
    """
    for ref in lista:
        try:
            setattr(sheet[ref], tipo, estilo)
        except Exception:
            pass  # ignora células que não existem


def header(ano, mes, moeda, smp):
    """
    Cria e formata o template inicial do Diário de Caixa em Excel.
    Retorna o workbook JÁ SALVO no disco.
    """
    wb = openpyxl.Workbook()
    sheet_name = f'Tabela Custo Entrada Saida {mes} {ano}'
    wb.create_sheet(sheet_name)
    wb.remove(wb['Sheet'])

    sheet = wb[sheet_name]

    # ____________________________________________________________
    # Labels
    # ____________________________________________________________
    sheet['F1'] = 'Ano:'
    sheet['F2'] = 'Mês:'
    sheet['F3'] = 'Moeda:'

    sheet['G1'] = ano
    sheet['G2'] = mes
    sheet['G3'] = moeda

    sheet['A5'] = 'DIÁRIO DE CAIXA'
    sheet.merge_cells('A5:G5')

    sheet['A6'] = 'Nº'
    sheet.merge_cells('A6:A7')
    sheet['B6'] = 'DATA'
    sheet.merge_cells('B6:B7')
    sheet['C6'] = 'DESIGNAÇÃO'
    sheet.merge_cells('C6:D7')
    sheet['E6'] = 'ENTRADAS (+)'
    sheet.merge_cells('E6:E7')
    sheet['F6'] = 'SAÍDAS (-)'
    sheet.merge_cells('F6:F7')
    sheet['G6'] = 'SALDO'
    sheet.merge_cells('G6:G7')

    sheet['D9'] = 'Saldo do Mês Anterior'
    sheet['G9'] = ' ' + str(smp)

    # ____________________________________________________________
    # Alinhamentos
    # ____________________________________________________________
    alinhar_centro = Alignment(horizontal='center', vertical='center')
    for ref in ['A5', 'A6', 'B6', 'C6', 'E6', 'F6', 'G6']:
        sheet[ref].alignment = alinhar_centro

    # ____________________________________________________________
    # Fontes
    # ____________________________________________________________
    fonte_preta = Font(name='Times New Roman', size=11, bold=True, color="000000")
    fonte_branca = Font(name='Times New Roman', size=11, bold=True, color="ffffff")

    for ref in ['F1', 'F2', 'F3', 'G1', 'G2', 'G3', 'A5']:
        sheet[ref].font = fonte_preta

    for ref in ['A6', 'B6', 'C6', 'E6', 'F6', 'G6', 'D9', 'E9', 'G9']:
        sheet[ref].font = fonte_branca

    # ____________________________________________________________
    # Cores de fundo
    # ____________________________________________________________
    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    brancos = ['A1', 'A2', 'A3', 'A4', 'A5', 'B1', 'B2', 'B3', 'B4',
               'C1', 'C2', 'C3', 'C4', 'D1', 'D2', 'D3', 'D4',
               'E1', 'E2', 'E3', 'E4', 'F1', 'F2', 'F3', 'F4',
               'G1', 'G2', 'G3', 'G4']
    cinzentos = ['A6', 'B6', 'C6', 'D6', 'D9', 'E6', 'E9', 'F6', 'F9', 'G6', 'G9']
    laranjas = ['A8', 'A9', 'B8', 'B9', 'C8', 'C9', 'D8', 'E8', 'F8', 'G8']

    estilizar(sheet, brancos, branco, 'fill')
    estilizar(sheet, cinzentos, cinzento, 'fill')
    estilizar(sheet, laranjas, laranja, 'fill')

    # ____________________________________________________________
    # Bordas
    # ____________________________________________________________
    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')

    borda_fina = openpyxl.styles.Border(
        left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina
    )
    borda_media = openpyxl.styles.Border(
        left=linha_media, right=linha_media, top=linha_media, bottom=linha_media
    )

    cabecalhos = ['A6', 'B6', 'C6', 'E6', 'F6', 'G6',
                   'A7', 'B7', 'C7', 'E7', 'F7', 'G7']
    saldo_celulas = ['D9', 'E9', 'F9', 'G9']

    estilizar(sheet, ['G1', 'G2', 'G3'], borda_fina, 'border')
    estilizar(sheet, cabecalhos, borda_media, 'border')
    estilizar(sheet, saldo_celulas,
              openpyxl.styles.Border(top=linha_media), 'border')
    estilizar(sheet, ['D9'],
              openpyxl.styles.Border(left=linha_media, top=linha_media), 'border')
    estilizar(sheet, ['G9'],
              openpyxl.styles.Border(left=linha_media, right=linha_media, top=linha_media), 'border')

    # ✅ Salvar o arquivo antes de retornar
    wb.save('Custo_Entrada&Saida.xlsx')
    return wb


def body(lista, mes, ano):
    """
    Adiciona os itens (transações) à planilha do Diário de Caixa.
    """
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')
    sheet_name = f'Tabela Custo Entrada Saida {mes} {ano}'
    sheet = wb[sheet_name]

    cont_table = 10  # primeira linha de dados

    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    borda_fina = openpyxl.styles.Border(
        left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina
    )
    fonte_preta = Font(name='Times New Roman', size=11, bold=True, color="000000")

    for item in lista:
        sheet.merge_cells(f'C{cont_table}:D{cont_table}')

        sheet[f'A{cont_table}'] = item[0]   # Nº
        sheet[f'B{cont_table}'] = item[1]   # Data
        sheet[f'C{cont_table}'] = item[2]   # Designação
        sheet[f'E{cont_table}'] = item[3]   # Entrada
        sheet[f'F{cont_table}'] = item[4]   # Saída
        sheet[f'G{cont_table}'] = item[5]   # Saldo

        for col in ('A', 'B', 'C', 'E', 'F', 'G'):
            sheet[f'{col}{cont_table}'].font = fonte_preta
            sheet[f'{col}{cont_table}'].border = borda_fina

        sheet[f'A{cont_table}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{cont_table}'].alignment = Alignment(horizontal='center', vertical='center')

        cont_table += 1

    # ✅ Salvar após adicionar os itens
    wb.save('Custo_Entrada&Saida.xlsx')
    return cont_table - 1


def footer(lista, mes, ano, sexo, nome, saldo_atual, p_total, saldo_anterior, despesa_anterior, ativos_total):
    """
    Adiciona o rodapé ao Diário de Caixa com totais e assinatura do responsável.
    """
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')
    sheet_name = f'Tabela Custo Entrada Saida {mes} {ano}'
    sheet = wb[sheet_name]

    tamanho = len(lista) + 10  # primeira linha do rodapé

    # ____________________________________________________________
    # Cores
    # ____________________________________________________________
    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    # ____________________________________________________________
    # Bordas
    # ____________________________________________________________
    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')
    borda_media = openpyxl.styles.Border(
        left=linha_media, right=linha_media, top=linha_media, bottom=linha_media
    )

    lista1 = [f'A{tamanho}', f'B{tamanho}', f'C{tamanho}', f'D{tamanho}',
              f'F{tamanho}', f'E{tamanho + 2}', f'F{tamanho + 2}',
              f'G{tamanho + 2}', f'G{tamanho + 4}']
    lista2 = [f'E{tamanho + 3}', f'F{tamanho + 3}', f'G{tamanho + 3}']

    # pintar fundos
    estilizar(sheet, lista1, cinzento, 'fill')
    estilizar(sheet, [f'E{tamanho}', f'G{tamanho}'], branco, 'fill')
    estilizar(sheet, [f'A{tamanho + 1}', f'B{tamanho + 1}', f'C{tamanho + 1}',
                      f'D{tamanho + 1}', f'E{tamanho + 1}', f'F{tamanho + 1}',
                      f'G{tamanho + 1}'], branco, 'fill')
    estilizar(sheet, [f'D{tamanho + 2}', f'D{tamanho + 3}', f'D{tamanho + 4}',
                      f'A{tamanho + 5}', f'A{tamanho + 6}', f'A{tamanho + 7}',
                      f'B{tamanho + 5}', f'B{tamanho + 6}', f'B{tamanho + 7}',
                      f'C{tamanho + 5}', f'C{tamanho + 6}', f'C{tamanho + 7}'], branco, 'fill')
    estilizar(sheet, lista2, laranja, 'fill')

    # bordas
    estilizar(sheet, lista1, borda_media, 'border')
    estilizar(sheet, lista2, borda_media, 'border')
    estilizar(sheet, [f'E{tamanho}', f'A{tamanho + 7}', f'B{tamanho + 7}', f'C{tamanho + 7}'],
              openpyxl.styles.Border(bottom=linha_fina), 'border')
    estilizar(sheet, [f'G{tamanho}'],
              openpyxl.styles.Border(right=linha_fina, bottom=linha_fina), 'border')

    # ____________________________________________________________
    # Preencher valores do rodapé
    # ____________________________________________________________
    sheet[f'D{tamanho + 2}'] = 'MOVIMENTOS DO MÊS'
    sheet[f'D{tamanho + 3}'] = 'SALDO DO MÊS ANTERIOR'
    sheet[f'D{tamanho + 4}'] = 'SALDO DO MÊS SEGUINTE'
    sheet[f'E{tamanho + 4}'] = '|------------»»»»»»»»»»»»'

    sheet[f'F{tamanho}'] = float(p_total)
    sheet[f'G{tamanho}'] = float(saldo_atual)
    sheet[f'E{tamanho + 2}'] = float(ativos_total)
    sheet[f'F{tamanho + 2}'] = float(p_total)
    sheet[f'G{tamanho + 2}'] = float(ativos_total - p_total)

    sheet[f'E{tamanho + 3}'] = float(saldo_anterior)
    sheet[f'F{tamanho + 3}'] = float(despesa_anterior)
    sheet[f'G{tamanho + 3}'] = float(saldo_anterior - despesa_anterior)

    sheet[f'G{tamanho + 4}'] = '+' + str(float(saldo_atual))

    # ____________________________________________________________
    # Fontes do rodapé
    # ____________________________________________________________
    fonte_branca = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    fonte_preta = Font(name='Times New Roman', size=11, bold=True, color="000000")
    fonte_normal = Font(name='Times New Roman', size=11, bold=False, color="000000")

    for ref in [f'E{tamanho + 2}', f'F{tamanho + 2}', f'G{tamanho + 2}',
                f'F{tamanho}', f'G{tamanho + 4}']:
        sheet[ref].font = fonte_branca

    for ref in [f'E{tamanho + 3}', f'F{tamanho + 3}', f'G{tamanho + 3}',
                f'G{tamanho}', f'D{tamanho + 2}', f'D{tamanho + 3}', f'D{tamanho + 4}']:
        sheet[ref].font = fonte_preta

    # ____________________________________________________________
    # Assinatura do responsável
    # ____________________________________________________________
    sheet.merge_cells(f'A{tamanho}:D{tamanho}')
    sheet.merge_cells(f'A{tamanho + 5}:B{tamanho + 5}')
    sheet.merge_cells(f'A{tamanho + 6}:B{tamanho + 6}')

    if sexo.upper() == 'F':
        sheet[f'A{tamanho + 5}'] = 'A TESOUREIRA'
    elif sexo.upper() == 'M':
        sheet[f'A{tamanho + 5}'] = 'O TESOUREIRO'
    else:
        sheet[f'A{tamanho + 5}'] = 'O/A TESOUREIRO/A'

    sheet[f'A{tamanho + 6}'] = str(nome)

    sheet[f'A{tamanho + 5}'].font = fonte_normal
    sheet[f'A{tamanho + 6}'].font = fonte_normal

    # ____________________________________________________________
    # Larguras das colunas
    # ____________________________________________________________
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 35
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 20

    # ✅ Salvar
    wb.save('Custo_Entrada&Saida.xlsx')


# =====================================================================
# PONTO DE ENTRADA
# =====================================================================
if __name__ == '__main__':
    if len(sys.argv) > 1:
        dados = json.loads(sys.argv[1])

        ano = str(dados.get('ano', ''))
        mes = str(dados.get('mes', '')).upper()
        moeda = str(dados.get('moeda', 'AKZ')).upper()
        saldo_anterior = float(dados.get('saldo_anterior', 0) or 0)
        despesa_anterior = float(dados.get('despesa_anterior', 0) or 0)
        sexo = str(dados.get('sexo', 'M'))
        nome = str(dados.get('nome', 'Administrador'))
        transacoes = dados.get('transacoes', [])

        # Saldo inicial = saldo anterior - despesa anterior
        saldo_inicial = saldo_anterior - despesa_anterior

        # 1. Criar o header (já salva o ficheiro)
        header(ano, mes, moeda, saldo_inicial)

        # 2. Processar as transações
        lista_itens = []
        saldo_atual = saldo_inicial
        p_total = 0.0
        ativos_total = 0.0

        for idx, trans in enumerate(transacoes, start=1):
            entrada = float(trans.get('entrada', 0) or 0)
            saida = float(trans.get('saida', 0) or 0)
            saldo_atual = saldo_atual + entrada - saida

            p_total += saida
            ativos_total += entrada

            item = [
                idx,
                trans.get('data', ''),
                trans.get('designacao', ''),
                entrada,
                saida,
                round(saldo_atual, 2)
            ]
            lista_itens.append(item)

        # 3. Adicionar body (se houver itens)
        if lista_itens:
            body(lista_itens, mes, ano)

        # 4. Adicionar footer
        footer(lista_itens, mes, ano, sexo, nome, saldo_atual,
               p_total, saldo_anterior, despesa_anterior, ativos_total)

        print('Custo_Entrada&Saida.xlsx')
    else:
        print('ERRO: Nenhum dado recebido')
        sys.exit(1)